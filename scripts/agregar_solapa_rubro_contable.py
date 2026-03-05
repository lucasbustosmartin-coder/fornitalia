#!/usr/bin/env python3
"""
Agrega al Excel Transacciones_Fornitalia una solapa con combinaciones únicas
Categoría + Cuenta contable, con Rubro contable sugerido y clasificación
Activo Corriente / Activo No Corriente / Pasivo.
Uso: python agregar_solapa_rubro_contable.py <ruta_al_excel>
"""

import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Se necesita openpyxl. Instalá con: pip install openpyxl")
    sys.exit(1)


# Reglas: (palabras clave en categoria/cuenta, rubro sugerido, tipo balance)
# Tipo: "Activo Corriente" | "Activo No Corriente" | "Pasivo"
REGLAS = [
    # Activo Corriente
    (["caja", "banco", "bancos", "cuenta corriente", "efectivo", "fondos"], "Caja y bancos", "Activo Corriente"),
    (["cliente", "clientes", "crédito por venta", "credito por venta", "cobranza", "cobro a cliente"], "Créditos por venta", "Activo Corriente"),
    (["inversión temporal", "inversion temporal", "plazo fijo", "caución", "cauciones", "fondo común"], "Inversiones temporarias", "Activo Corriente"),
    (["mercadería", "mercaderias", "bienes de cambio", "stock", "inventario"], "Bienes de cambio", "Activo Corriente"),
    (["deudor", "deudores", "anticipo", "anticipos a proveedor"], "Otros créditos (corriente)", "Activo Corriente"),
    # Activo No Corriente
    (["bien de uso", "bienes de uso", "activo fijo", "inmueble", "maquinaria", "vehículo", "vehiculo", "equipo"], "Bienes de uso", "Activo No Corriente"),
    (["inversión a largo", "inversion a largo", "inversión permanente"], "Inversiones (no corriente)", "Activo No Corriente"),
    (["intangibles", "marca", "software"], "Activos intangibles", "Activo No Corriente"),
    # Pasivo
    (["proveedor", "proveedores", "cuenta a pagar", "deuda con proveedor"], "Proveedores", "Pasivo"),
    (["préstamo", "prestamo", "financiación", "financiacion", "cuota préstamo", "cuota prestamo", "acreedor"], "Préstamos y deudas", "Pasivo"),
    (["impuesto", "iva", "ganancias", "ingresos brutos", "sellos", "afip", "arba", "fiscal"], "Impuestos y cargas sociales a pagar", "Pasivo"),
    (["sueldo", "sueldos", "salario", "cargas sociales", "personal a pagar", "honorario a pagar"], "Sueldos y cargas a pagar", "Pasivo"),
    (["venta", "ventas", "ingreso", "ingresos", "facturación", "facturacion", "cobranza (ingreso)"], "Ventas / Ingresos", "Pasivo"),
    (["alquiler a pagar", "locación a pagar", "locacion a pagar"], "Alquileres a pagar", "Pasivo"),
    (["comisión a pagar", "comision a pagar", "comisiones"], "Comisiones", "Pasivo"),
    (["seguro a pagar", "seguros a pagar"], "Seguros a pagar", "Pasivo"),
    # Gastos/resultado (cuentas de resultado se suelen presentar del lado pasivo/patrimonio)
    (["costo", "costo de venta", "insumo", "materia prima"], "Costo de ventas", "Pasivo"),
    (["gasto", "gastos", "administrativo", "servicio", "limpieza", "telefonía", "telefonia", "luz", "gas", "agua", "expensas"], "Gastos operativos", "Pasivo"),
    (["alquiler", "alquileres", "locación", "locacion", "inmueble (gasto)"], "Inmuebles / Alquileres", "Pasivo"),
    (["publicidad", "marketing", "promoción", "promocion"], "Publicidad y marketing", "Pasivo"),
    (["contador", "consultor", "asesor", "legal"], "Gastos administrativos", "Pasivo"),
    (["interés", "interes", "intereses", "gasto financiero"], "Intereses / Gastos financieros", "Pasivo"),
]


def normalizar(texto):
    if texto is None or (isinstance(texto, str) and not texto.strip()):
        return ""
    s = str(texto).strip().lower()
    # quitar acentos para matching
    for old, new in [("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"), ("ú", "u"), ("ñ", "n")]:
        s = s.replace(old, new)
    return s


def sugerir_rubro_y_tipo(categoria, cuenta_contable):
    c = normalizar(categoria)
    q = normalizar(cuenta_contable)
    texto = (c + " " + q).strip()
    if not texto:
        return "Sin clasificar", "Pasivo"
    for keys, rubro, tipo in REGLAS:
        if any(k in texto for k in keys):
            return rubro, tipo
    return "Otros gastos / Ingresos", "Pasivo"


def encontrar_columnas(ws):
    """Devuelve (fila del encabezado 1-based, idx_cat, idx_cuenta)."""
    for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        row_str = [str(c).strip().lower() if c is not None else "" for c in row]
        if "cuenta_contable" in row_str and "categoria" in row_str:
            return row_num, row_str.index("categoria") + 1, row_str.index("cuenta_contable") + 1
        if "categoria" in row_str:
            idx_cat = row_str.index("categoria") + 1
            for i, v in enumerate(row_str):
                if "cuenta" in v and "contable" in v:
                    return row_num, idx_cat, i + 1
        if "categoría" in row_str:
            idx_cat = row_str.index("categoría") + 1
            for i, v in enumerate(row_str):
                if "cuenta" in v and "contable" in v:
                    return row_num, idx_cat, i + 1
    return None, None, None


def main():
    if len(sys.argv) < 2:
        print("Uso: python scripts/agregar_solapa_rubro_contable.py <ruta_al_excel>")
        print("Ejemplo (desde la raíz del proyecto): python scripts/agregar_solapa_rubro_contable.py Transacciones_Fornitalia_27-02-2026.xlsx")
        sys.exit(1)
    path = Path(sys.argv[1])
    if not path.exists():
        print(f"No existe el archivo: {path}")
        sys.exit(1)
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        print("El archivo debe ser .xlsx o .xlsm")
        sys.exit(1)

    wb = load_workbook(path, read_only=False, data_only=True)
    ws0 = wb.worksheets[0]
    header_row, idx_cat, idx_cuenta = encontrar_columnas(ws0)
    if header_row is None or idx_cat is None or idx_cuenta is None:
        print("No se encontraron columnas 'categoria' y 'cuenta_contable' en la primera hoja.")
        sys.exit(1)

    # Recoger combinaciones únicas (categoria, cuenta_contable); datos desde la fila siguiente al encabezado
    seen = set()
    pares = []
    max_col = max(idx_cat, idx_cuenta)
    min_row_data = header_row + 1  # 1-based: primera fila de datos después del encabezado
    for row in ws0.iter_rows(min_row=min_row_data, min_col=1, max_col=max_col, values_only=True):
        row = list(row)
        while len(row) < max_col:
            row.append(None)
        cat_val = row[idx_cat - 1]
        cue_val = row[idx_cuenta - 1]
        cat = str(cat_val).strip() if cat_val is not None else ""
        cue = str(cue_val).strip() if cue_val is not None else ""
        if not cat:
            cat = "Sin categoría"
        if not cue:
            cue = "Sin cuenta"
        key = (cat, cue)
        if key in seen:
            continue
        seen.add(key)
        pares.append((cat, cue))

    pares.sort(key=lambda x: (x[0], x[1]))

    # Eliminar hoja "Rubro contable" si ya existe (para reemplazar)
    if "Rubro contable" in [s.title for s in wb.worksheets]:
        del wb["Rubro contable"]

    ws_rubro = wb.create_sheet("Rubro contable")
    ws_rubro.append(["Categoría", "Cuenta contable", "Rubro contable (sugerido)", "Tipo (Activo Corriente / Activo No Corriente / Pasivo)"])
    for cat, cue in pares:
        rubro, tipo = sugerir_rubro_y_tipo(cat, cue)
        ws_rubro.append([cat, cue, rubro, tipo])

    wb.save(path)
    print(f"Listo. Se agregó la solapa 'Rubro contable' al archivo: {path}")
    print(f"Combinaciones únicas: {len(pares)}.")


if __name__ == "__main__":
    main()
